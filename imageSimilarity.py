import openpyxl
import pandas as pd
import cv2
import numpy as np
import os
from keras.api.applications.vgg16 import VGG16, preprocess_input
from keras.api.preprocessing.image import img_to_array
from keras.api.models import Model
from skimage.metrics import structural_similarity as ssim

def orb_sim(img1, img2):
    orb = cv2.ORB_create()
    kp1, des1 = orb.detectAndCompute(img1, None)
    kp2, des2 = orb.detectAndCompute(img2, None)
    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
    matches = bf.match(des1, des2)
    similar_regions = [i for i in matches if i.distance < 59]
    if len(matches) == 0:
        return 0
    return len(similar_regions) / len(matches)

def structural_sim(img1, img2):
    sim, _ = ssim(img1, img2, full=True)
    return sim


def preprocess_image(image_path):
    img = cv2.imread(image_path)
    img = cv2.resize(img, (224, 224))
    img = img_to_array(img)
    img = np.expand_dims(img, axis=0)
    img = preprocess_input(img)
    return img

def vgg16_feature_similarity(image1, image2):
    img1 = preprocess_image(image1)
    img2 = preprocess_image(image2)
    feat1 = model.predict(img1)[0]
    feat2 = model.predict(img2)[0]
    similarity = np.dot(feat1, feat2) / (np.linalg.norm(feat1) * np.linalg.norm(feat2))
    return similarity

# VGG16 Modelini Yükle
base_model = VGG16(weights='imagenet')
model = Model(inputs=base_model.input, outputs=base_model.get_layer('fc1').output)

path = "dataCall.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

##cell_obj = sheet_obj.cell(row = 1, column = 1)

mx_row = sheet_obj.max_row

image_c1 = cv2.imread("c1.jpg", 0)
for i in range(2, mx_row + 1):
    temp = sheet_obj.cell(row = i, column = 3)
    comparison_image = cv2.imread(sheet_obj.cell(row = i, column = 1).value, 0)
    comparison_image_resized = cv2.resize(comparison_image, (image_c1.shape[1], image_c1.shape[0]))
    temp.value = orb_sim(image_c1, comparison_image_resized)
    print(temp.value)
    """test = sheet_obj.cell(row = i, column = 1)
    print(test.value)"""

wb_obj.save("dataCall.xlsx")

